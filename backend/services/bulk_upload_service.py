"""
JobRocket - Bulk Upload Service
Handles bulk job uploads from CSV and Excel files
Available for Pro and Enterprise tiers
"""

import io
import csv
import uuid
import logging
from datetime import datetime, timedelta
from typing import List, Dict, Any, Tuple, Optional
from motor.motor_asyncio import AsyncIOMotorDatabase

logger = logging.getLogger(__name__)


class BulkUploadService:
    """Service for bulk job uploads from CSV and Excel files"""
    
    REQUIRED_COLUMNS = ['title', 'description', 'location', 'salary', 'job_type']
    OPTIONAL_COLUMNS = ['work_type', 'industry', 'experience', 'qualifications', 'closing_date']
    
    VALID_JOB_TYPES = ['Permanent', 'Contract']
    VALID_WORK_TYPES = ['Remote', 'Onsite', 'Hybrid']
    
    def __init__(self, db: AsyncIOMotorDatabase):
        self.db = db
    
    async def process_file(
        self,
        file_content: bytes,
        filename: str,
        account_id: str,
        user_id: str,
        company_name: str,
        logo_url: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Process uploaded file and create jobs.
        
        Returns:
            {
                "success": bool,
                "total_rows": int,
                "created": int,
                "failed": int,
                "errors": [{"row": int, "error": str}]
            }
        """
        # Determine file type
        if filename.lower().endswith('.csv'):
            return await self._process_csv(
                file_content, account_id, user_id, company_name, logo_url
            )
        elif filename.lower().endswith(('.xlsx', '.xls')):
            return await self._process_excel(
                file_content, account_id, user_id, company_name, logo_url
            )
        else:
            return {
                "success": False,
                "total_rows": 0,
                "created": 0,
                "failed": 0,
                "errors": [{"row": 0, "error": f"Unsupported file format: {filename}. Use .csv or .xlsx"}]
            }
    
    async def _process_csv(
        self,
        content: bytes,
        account_id: str,
        user_id: str,
        company_name: str,
        logo_url: Optional[str]
    ) -> Dict[str, Any]:
        """Process CSV file"""
        try:
            # Decode and parse CSV
            text_content = content.decode('utf-8-sig')  # Handle BOM
            reader = csv.DictReader(io.StringIO(text_content))
            
            # Validate columns
            if reader.fieldnames is None:
                return {
                    "success": False,
                    "total_rows": 0,
                    "created": 0,
                    "failed": 0,
                    "errors": [{"row": 0, "error": "Could not read CSV headers"}]
                }
            
            # Normalize column names
            column_mapping = self._create_column_mapping(reader.fieldnames)
            missing_required = [col for col in self.REQUIRED_COLUMNS if col not in column_mapping]
            
            if missing_required:
                return {
                    "success": False,
                    "total_rows": 0,
                    "created": 0,
                    "failed": 0,
                    "errors": [{"row": 0, "error": f"Missing required columns: {', '.join(missing_required)}"}]
                }
            
            # Process rows
            rows = list(reader)
            return await self._process_rows(
                rows, column_mapping, account_id, user_id, company_name, logo_url
            )
            
        except UnicodeDecodeError:
            return {
                "success": False,
                "total_rows": 0,
                "created": 0,
                "failed": 0,
                "errors": [{"row": 0, "error": "Could not decode file. Please use UTF-8 encoding."}]
            }
        except Exception as e:
            logger.error(f"Error processing CSV: {e}")
            return {
                "success": False,
                "total_rows": 0,
                "created": 0,
                "failed": 0,
                "errors": [{"row": 0, "error": f"Error processing file: {str(e)}"}]
            }
    
    async def _process_excel(
        self,
        content: bytes,
        account_id: str,
        user_id: str,
        company_name: str,
        logo_url: Optional[str]
    ) -> Dict[str, Any]:
        """Process Excel file"""
        try:
            import openpyxl
            
            # Load workbook
            workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            sheet = workbook.active
            
            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value) if cell.value else '')
            
            # Validate columns
            column_mapping = self._create_column_mapping(headers)
            missing_required = [col for col in self.REQUIRED_COLUMNS if col not in column_mapping]
            
            if missing_required:
                return {
                    "success": False,
                    "total_rows": 0,
                    "created": 0,
                    "failed": 0,
                    "errors": [{"row": 0, "error": f"Missing required columns: {', '.join(missing_required)}"}]
                }
            
            # Convert rows to dictionaries
            rows = []
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if all(cell is None or cell == '' for cell in row):
                    continue  # Skip empty rows
                row_dict = {}
                for i, header in enumerate(headers):
                    if i < len(row):
                        row_dict[header] = row[i]
                rows.append(row_dict)
            
            workbook.close()
            
            return await self._process_rows(
                rows, column_mapping, account_id, user_id, company_name, logo_url
            )
            
        except ImportError:
            return {
                "success": False,
                "total_rows": 0,
                "created": 0,
                "failed": 0,
                "errors": [{"row": 0, "error": "Excel support not installed. Please contact support."}]
            }
        except Exception as e:
            logger.error(f"Error processing Excel: {e}")
            return {
                "success": False,
                "total_rows": 0,
                "created": 0,
                "failed": 0,
                "errors": [{"row": 0, "error": f"Error processing file: {str(e)}"}]
            }
    
    def _create_column_mapping(self, headers: List[str]) -> Dict[str, str]:
        """Create mapping from normalized column names to actual headers"""
        mapping = {}
        
        # Common variations
        variations = {
            'title': ['title', 'job_title', 'job title', 'position', 'role'],
            'description': ['description', 'job_description', 'job description', 'details', 'summary'],
            'location': ['location', 'city', 'area', 'place', 'office_location'],
            'salary': ['salary', 'compensation', 'pay', 'remuneration', 'salary_range'],
            'job_type': ['job_type', 'job type', 'type', 'employment_type', 'employment type'],
            'work_type': ['work_type', 'work type', 'remote', 'work_mode', 'work mode'],
            'industry': ['industry', 'sector', 'category', 'field'],
            'experience': ['experience', 'experience_required', 'years', 'exp'],
            'qualifications': ['qualifications', 'requirements', 'skills', 'education'],
            'closing_date': ['closing_date', 'closing date', 'deadline', 'expires', 'end_date']
        }
        
        for header in headers:
            header_lower = header.lower().strip()
            for standard_name, aliases in variations.items():
                if header_lower in aliases:
                    mapping[standard_name] = header
                    break
        
        return mapping
    
    async def _process_rows(
        self,
        rows: List[Dict],
        column_mapping: Dict[str, str],
        account_id: str,
        user_id: str,
        company_name: str,
        logo_url: Optional[str]
    ) -> Dict[str, Any]:
        """Process validated rows and create jobs"""
        total_rows = len(rows)
        created = 0
        failed = 0
        errors = []
        
        for row_num, row in enumerate(rows, start=2):  # Start at 2 (header is row 1)
            try:
                # Extract values using column mapping
                def get_value(key: str, default: Any = None) -> Any:
                    if key in column_mapping:
                        return row.get(column_mapping[key], default) or default
                    return default
                
                title = get_value('title')
                description = get_value('description')
                location = get_value('location')
                salary = get_value('salary')
                job_type = get_value('job_type', 'Permanent')
                work_type = get_value('work_type', 'Onsite')
                industry = get_value('industry', 'General')
                experience = get_value('experience')
                qualifications = get_value('qualifications')
                closing_date_str = get_value('closing_date')
                
                # Validate required fields
                if not title:
                    errors.append({"row": row_num, "error": "Missing title"})
                    failed += 1
                    continue
                
                if not description:
                    errors.append({"row": row_num, "error": "Missing description"})
                    failed += 1
                    continue
                
                if not location:
                    errors.append({"row": row_num, "error": "Missing location"})
                    failed += 1
                    continue
                
                if not salary:
                    errors.append({"row": row_num, "error": "Missing salary"})
                    failed += 1
                    continue
                
                # Validate job_type
                job_type_normalized = job_type.capitalize() if job_type else 'Permanent'
                if job_type_normalized not in self.VALID_JOB_TYPES:
                    job_type_normalized = 'Permanent'
                
                # Validate work_type
                work_type_normalized = work_type.capitalize() if work_type else 'Onsite'
                if work_type_normalized not in self.VALID_WORK_TYPES:
                    work_type_normalized = 'Onsite'
                
                # Parse closing date
                closing_date = None
                if closing_date_str:
                    try:
                        if isinstance(closing_date_str, datetime):
                            closing_date = closing_date_str
                        else:
                            # Try common date formats
                            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y']:
                                try:
                                    closing_date = datetime.strptime(str(closing_date_str), fmt)
                                    break
                                except:
                                    continue
                    except:
                        pass
                
                # Create job document
                job_dict = {
                    "id": str(uuid.uuid4()),
                    "account_id": account_id,
                    "posted_by": user_id,
                    "company_name": company_name,
                    "logo_url": logo_url,
                    "title": str(title).strip(),
                    "description": str(description).strip(),
                    "location": str(location).strip(),
                    "salary": str(salary).strip(),
                    "job_type": job_type_normalized,
                    "work_type": work_type_normalized,
                    "industry": str(industry).strip() if industry else "General",
                    "experience": str(experience).strip() if experience else None,
                    "qualifications": str(qualifications).strip() if qualifications else None,
                    "closing_date": closing_date,
                    "posted_date": datetime.utcnow(),
                    "expiry_date": datetime.utcnow() + timedelta(days=35),
                    "is_active": True,
                    "featured": False,
                    "source": "bulk_upload"
                }
                
                await self.db.jobs.insert_one(job_dict)
                created += 1
                
            except Exception as e:
                logger.error(f"Error creating job from row {row_num}: {e}")
                errors.append({"row": row_num, "error": str(e)})
                failed += 1
        
        return {
            "success": failed == 0,
            "total_rows": total_rows,
            "created": created,
            "failed": failed,
            "errors": errors[:50]  # Limit error list to 50
        }
    
    def generate_template(self, format: str = 'csv') -> Tuple[bytes, str]:
        """
        Generate a template file for bulk upload.
        
        Returns:
            (content_bytes, filename)
        """
        headers = self.REQUIRED_COLUMNS + self.OPTIONAL_COLUMNS
        
        sample_row = {
            'title': 'Software Developer',
            'description': 'We are looking for a skilled developer to join our team...',
            'location': 'Johannesburg, Gauteng',
            'salary': 'R50,000 - R70,000 per month',
            'job_type': 'Permanent',
            'work_type': 'Hybrid',
            'industry': 'Technology',
            'experience': '3+ years',
            'qualifications': 'BSc Computer Science or equivalent',
            'closing_date': '2025-03-31'
        }
        
        if format == 'csv':
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=headers)
            writer.writeheader()
            writer.writerow(sample_row)
            return output.getvalue().encode('utf-8'), 'job_upload_template.csv'
        
        elif format == 'xlsx':
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = 'Jobs'
            
            # Header styling
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
            
            # Write headers
            for col, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
            
            # Write sample row
            for col, header in enumerate(headers, start=1):
                sheet.cell(row=2, column=col, value=sample_row.get(header, ''))
            
            # Save to bytes
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            return output.read(), 'job_upload_template.xlsx'
        
        raise ValueError(f"Unsupported format: {format}")


def create_bulk_upload_service(db: AsyncIOMotorDatabase) -> BulkUploadService:
    """Factory function to create bulk upload service"""
    return BulkUploadService(db)
